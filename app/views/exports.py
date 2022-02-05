from django.http.response import JsonResponse
from django.shortcuts import redirect, render, HttpResponse
from django.views import View
from ..forms import *
from rest_framework.views import APIView
from ..serializers import *
from ..models import *
import sweetify
from datetime import date as now
from decimal import Decimal
from datetime import datetime
from django.core.exceptions import PermissionDenied
from .journalAPI import jeAPI
import re
import openpyxl as op
from .notificationCreate import *


########## EXPORTS ##########
class ExportsView(View):
    def get(self, request):
        if request.user.authLevel == '2':
            raise PermissionDenied()

        user = request.user

        try:
            exports = user.branch.exports.latest('pk')

            listed_code = exports.code.split('-')
            listed_date = str(now.today()).split('-')

            current_code = int(listed_code[3])

            if listed_code[1] == listed_date[0] and listed_code[2] == listed_date[1]:
                current_code += 1
                new_code = 'EXP-{}-{}-{}'.format(listed_date[0], listed_date[1], str(current_code).zfill(4))
            else:
                new_code = 'EXP-{}-{}-0001'.format(listed_date[0], listed_date[1])

        except Exception as e:
            print(e)
            listed_date = str(now.today()).split('-')
            new_code = 'EXP-{}-{}-0001'.format(listed_date[0], listed_date[1])

        context = {
            'new_code': new_code,
        }
        return render(request, 'exports.html', context)

class ExportsListView(View):
    def get(self, request, format=None):
        if request.user.authLevel == '2':
            raise PermissionDenied()
        return render(request, 'exports-list.html')

class SaveExports(APIView):
    def post(self, request, format = None):
        if request.user.authLevel == '2':
            raise PermissionDenied()
        export = request.data

        exp = Exports()

        exp.code = export['code']
        exp.datetimeCreated = datetime.now()

        if export['retroactive']:
            exp.dateSold = export['retroactive']
        else:
            exp.dateSold = export['date']

        exp.party = Party.objects.get(pk=export['customer'])
        
        exp.amountDue = Decimal(export['amountDue'])
        exp.amountTotal = Decimal(export['amountTotal'])
        exp.runningBalance = exp.amountTotal
        exp.chequeNo = export['chequeNo']
        exp.dueDate = export['dueDate']
        exp.bank = export['bank']
        exp.remarks = export['remarks']
        exp.forex = export['exchangeRate']
        exp.runningBalancePeso = Decimal(exp.amountTotal) * Decimal(exp.forex)

        if request.user.is_authenticated:
            exp.createdBy = request.user

        if export['discountType'] == 'percent':
            exp.discountPercent = export['totalDiscount']
        elif export['discountType'] == 'peso':
            exp.discountPeso = export['totalDiscount']

        exp.save()
        request.user.branch.exports.add(exp)

        

        for item in export['items']:
            exportitemsmerch = ExportItemsMerch()
            exportitemsmerch.export = exp
            exportitemsmerch.merchInventory = MerchandiseInventory.objects.get(pk=item['code'])
            exportitemsmerch.remaining = item['remaining']
            exportitemsmerch.qty = item['quantity']
            exportitemsmerch.pallet = item['pallet']
            exportitemsmerch.cbm = item['cbm']
            exportitemsmerch.vol = item['vol']
            exportitemsmerch.pricePerCubic = Decimal(item['pricePerCubic'])
            exportitemsmerch.totalCost = Decimal(item['totalCost'])

            print(item['pricePerCubic'], item['totalCost'])
            
            exportitemsmerch.save()
            request.user.branch.exportItemsMerch.add(exportitemsmerch)

        for fee in export['otherFees']:
            f = ExportOtherFees()
            f.export = exp
            f.fee = fee['fee']
            f.description = fee['description']
            f.save()
            request.user.branch.exportOtherFees.add(f)

        notify(request, 'New Exports Request', exp.code, '/exports-nonapproved/', 1)
        sweetify.sweetalert(request, icon='success', title='Success!', persistent='Dismiss')
        return JsonResponse(0, safe=False)


class ReceivedPaymentsUSDView(View):
    def get(self, request):
        if request.user.authLevel == '2':
            raise PermissionDenied()
        
        try:
            rp = request.user.branch.receivePayment.latest('pk')

            listed_code = rp.code.split('-')
            listed_date = str(now.today()).split('-')

            current_code = int(listed_code[3])

            if listed_code[1] == listed_date[0] and listed_code[2] == listed_date[1]:
                current_code += 1
                new_code = 'RP/USD-{}-{}-{}'.format(listed_date[0], listed_date[1], str(current_code).zfill(4))
            else:
                new_code = 'RP/USD-{}-{}-0001'.format(listed_date[0], listed_date[1])

        except Exception as e:
            print(e)
            listed_date = str(now.today()).split('-')
            new_code = 'RP/USD-{}-{}-0001'.format(listed_date[0], listed_date[1])

        context = {
            'new_code': new_code,
            'exports': request.user.branch.exports.filter(approved=True, fullyPaid=False).exclude(runningBalance=Decimal(0)),
            'customers': request.user.branch.party.filter(type="Customer")
        }
        return render(request, 'received-payments-usd.html', context)

class SaveReceivePaymentsUSD(APIView):
    def post(self, request, format = None):
        if request.user.authLevel == '2':
            raise PermissionDenied()
        dChildAccount = request.user.branch.branchProfile.branchDefaultChildAccount
        receivePaymentUSD = request.data


        ##### RECEIVED PAYMENT DEFAULT #####
        if receivePaymentUSD['rpType'] == 'Default':

            rp = ReceivePaymentUSD()
            rp.code = receivePaymentUSD['code']
            rp.datetimeCreated = datetime.now()
            rp.remarks = receivePaymentUSD['description']
            if receivePaymentUSD['retroactive']:
                rp.paymentDate = receivePaymentUSD['retroactive']
            else:
                rp.paymentDate = receivePaymentUSD['date']
            if request.user.is_authenticated:
                rp.createdBy = request.user
            rp.exports = Exports.objects.get(pk=receivePaymentUSD['sc']['code'])
            rp.paymentMethod = receivePaymentUSD['paymentMethod']
            rp.paymentPeriod = receivePaymentUSD['paymentPeriod']
            rp.forex = Decimal(receivePaymentUSD['exchangeRate'])

            rp.amountPaid = Decimal(receivePaymentUSD['amountPaid'])
            rp.save()
            request.user.branch.receivepaymentsUSD.add(rp)

            if rp.paymentMethod == "Cheque":
                cheque = Cheques()
                cheque.chequeNo = receivePaymentUSD['chequeNo']
                cheque.accountChild = AccountChild.objects.get(pk=receivePaymentUSD['bank'])
                print(receivePaymentUSD['bank'])
                cheque.dueDate = receivePaymentUSD['dueDate']
                cheque.save()
                request.user.branch.cheque.add(cheque)

                rp.cheque = cheque
                rp.save()

                sweetify.sweetalert(request, icon='success', title='Success!', persistent='Dismiss')
                return JsonResponse(0, safe=False)
                
            

            j = Journal()
            j.code = rp.code
            j.datetimeCreated = rp.datetimeCreated
            j.createdBy = rp.createdBy
            j.journalDate = datetime.now()
            j.save()
            rp.journal = j
            rp.save()
            request.user.branch.journal.add(j)

            rp.exports.runningBalance -= (rp.amountPaid)
            rp.exports.runningBalancePeso -= (rp.amountPaid*rp.forex)
            if rp.exports.runningBalance == 0:
                rp.exports.fullyPaid = True

            if rp.paymentPeriod == "Partial Payment":
                jeAPI(request, j, 'Credit', rp.exports.party.accountChild.get(name__regex=r"[Rr]eceivable"), (rp.amountPaid*rp.forex))
                jeAPI(request, j, 'Debit', dChildAccount.cashInBank.get(name=rp.paymentMethod), rp.amountPaid*rp.forex)
            elif rp.paymentPeriod == "Full Payment":
                if rp.exports.runningBalancePeso < 0:
                    jeAPI(request, j, 'Credit', rp.exports.party.accountChild.get(name__regex=r"[Rr]eceivable"), (rp.amountPaid*rp.forex))
                    jeAPI(request, j, 'Credit', dChildAccount.forexGain, abs(rp.exports.runningBalancePeso))
                    jeAPI(request, j, 'Debit', dChildAccount.cashInBank.get(name=rp.paymentMethod), rp.amountPaid*rp.forex + abs(rp.exports.runningBalancePeso))
                elif rp.exports.runningBalancePeso > 0:
                    jeAPI(request, j, 'Credit', rp.exports.party.accountChild.get(name__regex=r"[Rr]eceivable"), (rp.amountPaid*rp.forex))
                    jeAPI(request, j, 'Debit', dChildAccount.forexLoss, abs(rp.exports.runningBalancePeso))
                    jeAPI(request, j, 'Debit', dChildAccount.cashInBank.get(name=rp.paymentMethod), rp.amountPaid*rp.forex - abs(rp.exports.runningBalancePeso))
                elif rp.exports.runningBalancePeso== 0:
                    jeAPI(request, j, 'Credit', rp.exports.party.accountChild.get(name__regex=r"[Rr]eceivable"), (rp.amountPaid*rp.forex))
                    jeAPI(request, j, 'Debit', dChildAccount.cashInBank.get(name=rp.paymentMethod), rp.amountPaid*rp.forex)
            

            rp.exports.save()

        sweetify.sweetalert(request, icon='success', title='Success!', persistent='Dismiss')
        return JsonResponse(0, safe=False)

class ExportCompercialInvoice(APIView):
    def get(self, request, pk):
        expo = Exports.objects.get(pk=pk)

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="Commercial Invoice {expo.code}.xlsx"'

        wb = op.load_workbook('static/files/commercialInvoice.xlsx')
        ws = wb.active  

        rpdate = ws.cell(row=5, column=7)
        rpcode = ws.cell(row=6, column=7)
        partyName = ws.cell(row=12, column=1)
        partyAddress = ws.cell(row=13, column=1)
        contactNum = ws.cell(row=14, column=1)
        contactPerson = ws.cell(row=15, column=1)
        remarks = ws.cell(row=12, column=7)

        totalAmount = ws.cell(row=36, column=8)
        totalQty = ws.cell(row=36, column=4)
        totalPallet = ws.cell(row=36, column=5)
        totalVol = ws.cell(row=36, column=6)

        rpdate.value = expo.receivepaymentUSD.earliest('pk').datetimeCreated
        rpcode.value = expo.receivepaymentUSD.earliest('pk').code
        partyName.value = expo.party.name
        partyAddress.value = expo.party.shippingAddress if expo.party.shippingAddress else expo.party.officeAddress
        contactNum.value = f"Tel.: {expo.party.landline} / Fax:"
        contactPerson.value = f"Contact Persion: {expo.party.contactPerson}"
        remarks.value = expo.remarks
        totalAmount.value = expo.amountTotal

        ctr = 0
        qty=0
        pallet=0
        vol=0
        for i in expo.exportitemsmerch.all():
            invName = ws.cell(row=21+ctr, column=1)
            size = ws.cell(row=21+ctr, column=3)
            qtyPCS = ws.cell(row=21+ctr, column=4)
            qtyPallet = ws.cell(row=21+ctr, column=5)
            cubic = ws.cell(row=21+ctr, column=6)
            cubicPrice = ws.cell(row=21+ctr, column=7)
            amount = ws.cell(row=21+ctr, column=8)

            invName.value = f"{i.merchInventory.name} {i.merchInventory.classification}"
            size.value = f"{round(i.merchInventory.thickness, 0)} x {round(i.merchInventory.width, 0)} x {round(i.merchInventory.length, 0)}"
            qtyPCS.value = i.qty
            qtyPallet.value = i.pallet
            cubic.value = i.vol
            cubicPrice.value = i.pricePerCubic
            amount.value = i.totalCost
            qty += i.qty
            pallet += i.pallet
            vol += i.vol

            ctr+=1

        totalQty.value = qty
        totalPallet.value = pallet
        totalVol.value = vol
        
        wb.save(response)
        return response
            



class ExportPackingList(APIView):
    def get(self, request, pk):
        expo = Exports.objects.get(pk=pk)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="Packing List {expo.code}.xlsx"'

        wb = op.load_workbook('static/files/packingList.xlsx')
        ws = wb.active

        rpdate = ws.cell(row=5, column=6)
        rpcode = ws.cell(row=6, column=6)

        partyName = ws.cell(row=10, column=1)
        partyAddress = ws.cell(row=11, column=1)
        contactNum = ws.cell(row=12, column=1)
        contactPerson = ws.cell(row=13, column=1)
        remarks = ws.cell(row=8, column=6)

        rpdate.value = expo.receivepaymentUSD.earliest('pk').datetimeCreated
        rpcode.value = expo.receivepaymentUSD.earliest('pk').code

        partyName.value = expo.party.name
        partyAddress.value = expo.party.shippingAddress if expo.party.shippingAddress else expo.party.officeAddress
        contactNum.value = f"Tel.: {expo.party.landline} / Fax:"
        contactPerson.value = f"Contact Persion: {expo.party.contactPerson}"
        remarks.value = expo.remarks

        totalQty = ws.cell(row=38, column=4)
        totalPallet = ws.cell(row=38, column=5)

        totalQty.value = 0
        totalPallet.value = 0

        ctr = 0
        for i in expo.exportitemsmerch.all():
            invName = ws.cell(row=19+ctr, column=1)
            palletNo = ws.cell(row=19+ctr, column=2)
            size = ws.cell(row=19+ctr, column=3)
            qtyPCS = ws.cell(row=19+ctr, column=4)
            qtyPallet = ws.cell(row=19+ctr, column=5)

            invName.value = f"{i.merchInventory.name} {i.merchInventory.classification}"
            palletNo.value = ctr+1
            size.value = f"{round(i.merchInventory.thickness, 0)} x {round(i.merchInventory.width, 0)} x {round(i.merchInventory.length, 0)}"
            qtyPCS.value = i.qty
            qtyPallet.value = i.pallet

            totalQty.value += i.qty
            totalPallet.value += i.pallet

            ctr+=1
            
        wb.save(response)
        return response