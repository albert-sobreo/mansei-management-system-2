from django.http.response import JsonResponse
from django.shortcuts import redirect, render, HttpResponse
from django.views import View
from ..forms import *
from rest_framework.views import APIView
from ..serializers import *
from ..models import *
import sweetify
from datetime import date as now
from decimal import Decimal
from datetime import datetime
from django.core.exceptions import PermissionDenied
import pandas as pd
import openpyxl as op
from .notificationCreate import *

########## SALES CONTRACT ##########
class SalesContractView(View):
    def get(self, request, format=None):
        if request.user.authLevel == '2':
            raise PermissionDenied()

        user = request.user

        try:
            sc = user.branch.salesContract.latest('pk')

            listed_code = sc.code.split('-')
            listed_date = str(now.today()).split('-')

            current_code = int(listed_code[3])

            if listed_code[1] == listed_date[0] and listed_code[2] == listed_date[1]:
                current_code += 1
                new_code = 'SC-{}-{}-{}'.format(listed_date[0], listed_date[1], str(current_code).zfill(4))
            else:
                new_code = 'SC-{}-{}-0001'.format(listed_date[0], listed_date[1])

        except Exception as e:
            print(e)
            listed_date = str(now.today()).split('-')
            new_code = 'SC-{}-{}-0001'.format(listed_date[0], listed_date[1])

        context = {
            'new_code': new_code,
            "salesOrder": request.user.branch.salesOrder.all().filter(approved=True,soed=False)
        }
        return render(request, 'sales-contract.html', context)

class SCListView(View):
    def get(self, request, format=None):
        if request.user.authLevel == '2':
            raise PermissionDenied()
        return render(request, 'sc-list.html')

class SaveSalesContract(APIView):
    def post(self, request, format = None):
        if request.user.authLevel == '2':
            raise PermissionDenied()
        salesContract = request.data

        sc = SalesContract()

        sc.code = salesContract['code']
        sc.datetimeCreated = datetime.now()

        if salesContract['retroactive']:
            sc.dateSold = salesContract['retroactive']
        else:
            sc.dateSold = salesContract['date']

        sc.party = Party.objects.get(pk=salesContract['customer'])
        
        sc.amountDue = Decimal(salesContract['amountDue'])
        sc.amountTotal = Decimal(salesContract['amountTotal'])
        sc.runningBalance = sc.amountTotal
        sc.taxType = salesContract['taxType']
        sc.taxRate = Decimal(salesContract['taxRate'])
        sc.taxPeso = Decimal(salesContract['taxPeso'])
        sc.chequeNo = salesContract['chequeNo']
        sc.dueDate = salesContract['dueDate']
        sc.bank = salesContract['bank']
        sc.remarks = salesContract['remarks']
        sc.wep = Decimal(salesContract['withholdingTax'])

        if salesContract['salesOrder']:
            sc.salesOrder = SalesOrder.objects.get(pk=salesContract['salesOrder'])
            sc.salesOrder.soed = True
            sc.salesOrder.save()

        if request.user.is_authenticated:
            sc.createdBy = request.user

        if salesContract['discountType'] == 'percent':
            sc.discountPercent = salesContract['totalDiscount']
        elif salesContract['discountType'] == 'peso':
            sc.discountPeso = salesContract['totalDiscount']

        sc.save()
        request.user.branch.salesContract.add(sc)


        atc = SCatc()

        for jsonatc in salesContract['atc']:
            print(jsonatc)
            atc.code = ATC.objects.get(pk=jsonatc['code'])
            atc.amountWithheld = jsonatc['amountWithheld']
            atc.salesContract = sc
            atc.save()
            request.user.branch.scatc.add(atc)
        

        for item in salesContract['items']:
            scitemsmerch = SCItemsMerch()
            scitemsmerch.salesContract = sc
            scitemsmerch.merchInventory = MerchandiseInventory.objects.get(pk=item['code'])
            scitemsmerch.remaining = item['remaining']
            scitemsmerch.qty = item['quantity']
            scitemsmerch.cbm = item['cbm']
            scitemsmerch.vol = item['vol']
            scitemsmerch.pricePerCubic = Decimal(item['pricePerCubic'])
            scitemsmerch.totalCost = Decimal(item['totalCost'])

            print(item['pricePerCubic'], item['totalCost'])
            
            scitemsmerch.save()
            request.user.branch.scitemsMerch.add(scitemsmerch)

        for fee in salesContract['otherFees']:
            f = TempSCOtherFees()
            f.salesContract = sc
            f.fee = fee['fee']
            f.description = fee['description']
            f.save()
            request.user.branch.tempSCOtherFees.add(f)

        notify(request, 'New SC Request', sc.code, '/sc-nonapproved/', 1)
        sweetify.sweetalert(request, icon='success', title='Success!', persistent='Dismiss')
        return JsonResponse(0, safe=False)






########## QUOTATIONS ##########
class SalesQuotationsView(View):
    def get(self, request, format=None):
        if request.user.authLevel == '2':
            raise PermissionDenied()

        user = request.user

        try:
            qq = user.branch.quotations.latest('pk')

            listed_code = qq.code.split('-')
            listed_date = str(now.today()).split('-')

            current_code = int(listed_code[3])

            if listed_code[1] == listed_date[0] and listed_code[2] == listed_date[1]:
                current_code += 1
                new_code = 'QQ-{}-{}-{}'.format(listed_date[0], listed_date[1], str(current_code).zfill(4))
            else:
                new_code = 'QQ-{}-{}-0001'.format(listed_date[0], listed_date[1])

        except Exception as e:
            print(e)
            listed_date = str(now.today()).split('-')
            new_code = 'QQ-{}-{}-0001'.format(listed_date[0], listed_date[1])

        context = {
            'new_code': new_code,
        }
        return render(request, 'sales-quotations.html', context)

class QQListView(View):
    def get(self, request, format=None):
        if request.user.authLevel == '2':
            raise PermissionDenied()
        return render(request, 'qq-list.html')

class SaveQuotations(APIView):
    def post(self, request, format = None):
        if request.user.authLevel == '2':
            raise PermissionDenied()
        quotes = request.data

        qq = Quotations()

        qq.code = quotes['code']
        qq.datetimeCreated = datetime.now()
        qq.dateQuoted = quotes['date']
        qq.party = Party.objects.get(pk=quotes['customer'])
        qq.amountDue = Decimal(quotes['amountDue'])
        qq.amountTotal = Decimal(quotes['amountTotal'])
        qq.taxType = quotes['taxType']
        qq.taxRate = Decimal(quotes['taxRate'])
        qq.taxPeso = Decimal(quotes['taxPeso'])
        qq.wep = Decimal(quotes['withholdingTax'])

        if request.user.is_authenticated:
            qq.createdBy = request.user

        if quotes['discountType'] == 'percent':
            qq.discountPercent = quotes['totalDiscount']
        elif quotes['discountType'] == 'peso':
            qq.discountPeso = quotes['totalDiscount']

        qq.save()
        request.user.branch.quotations.add(qq)
        
        atc = QQatc()
        
        for jsonatc in quotes['atc']:
            print(jsonatc)
            atc.code = ATC.objects.get(pk=jsonatc['code'])
            atc.amountWithheld = jsonatc['amountWithheld']
            atc.quotations = qq
            atc.save()
            request.user.branch.qqatc.add(atc)

        for item in quotes['items']:
            qqitemsmerch = QQItemsMerch()
            qqitemsmerch.quotations = qq
            qqitemsmerch.pricePerCubic = item['pricePerCubic']
            qqitemsmerch.merchInventory = MerchandiseInventory.objects.get(pk=item['code'])
            qqitemsmerch.warehouse = qqitemsmerch.merchInventory.warehouseitems.all()[0].warehouse
            qqitemsmerch.remaining = item['remaining']
            qqitemsmerch.qty = item['quantity']
            qqitemsmerch.totalCost = Decimal(item['totalCost'])
            qqitemsmerch.cbm = item['cbm']
            qqitemsmerch.vol = item['vol']


            print(qqitemsmerch.totalCost)
            
            qqitemsmerch.save()
            request.user.branch.qqitemsMerch.add(qqitemsmerch)

        for fee in quotes['otherFees']:
            f = QQCOtherFees()
            f.quotations = qq
            f.fee = fee['fee']
            f.description = fee['description']
            f.save()
            request.user.branch.qqOtherFees.add(f)

        notify(request, 'New Quotation Request', qq.code, '/qq-nonapproved/', 1)
        sweetify.sweetalert(request, icon='success', title='Success!', persistent='Dismiss')
        return JsonResponse(0, safe=False)










########## SALES ORDER ##########
class SalesOrderView(View):
    def get(self, request, format=None):
        if request.user.authLevel == '2':
            raise PermissionDenied()
        
        try:
            so = request.user.branch.salesOrder.latest('pk')

            listed_code = so.code.split('-')
            listed_date = str(now.today()).split('-')

            current_code = int(listed_code[3])

            if listed_code[1] == listed_date[0] and listed_code[2] == listed_date[1]:
                current_code += 1
                new_code = 'SO-{}-{}-{}'.format(listed_date[0], listed_date[1], str(current_code).zfill(4))
            else:
                new_code = 'SO-{}-{}-0001'.format(listed_date[0], listed_date[1])

        except Exception as e:
            print(e)
            listed_date = str(now.today()).split('-')
            new_code = 'SO-{}-{}-0001'.format(listed_date[0], listed_date[1])

        print(SalesOrder.objects.none())

        context = {
            'new_code': new_code,
            "quotations": request.user.branch.quotations.all().filter(approved=True, qqd=False)
        }
        return render(request, 'sales-order.html', context)

class SOListView(View):
    def get(self, request, format=None):
        if request.user.authLevel == '2':
            raise PermissionDenied()
        return render(request, 'so-list.html')

class SaveSalesOrder(APIView):
    def post(self, request, format = None):
        if request.user.authLevel == '2':
            raise PermissionDenied()
        salesOrder = request.data

        so = SalesOrder()

        so.code = salesOrder['code']
        so.datetimeCreated = datetime.now()
        
        if salesOrder['retroactive']:
            so.dateSold = salesOrder['retroactive']
        else:
            so.dateSold = salesOrder['date']

        so.party = Party.objects.get(pk=salesOrder['customer'])

        if salesOrder['quotation']:
            so.quotations = Quotations.objects.get(pk=salesOrder['quotation'])
            so.quotations.qqd = True
            so.quotations.save()

        so.amountDue = Decimal(salesOrder['amountDue'])
        so.amountTotal = Decimal(salesOrder['amountTotal'])
        so.taxType = salesOrder['taxType']
        so.taxRate = Decimal(salesOrder['taxRate'])
        so.taxPeso = Decimal(salesOrder['taxPeso'])
        so.chequeNo = salesOrder['chequeNo']
        so.dueDate = salesOrder['dueDate']
        so.bank = salesOrder['bank']
        so.remarks = salesOrder['remarks']

        if request.user.is_authenticated:
            so.createdBy = request.user

        if salesOrder['discountType'] == 'percent':
            so.discountPercent = salesOrder['totalDiscount']
        elif salesOrder['discountType'] == 'peso':
            so.discountPeso = salesOrder['totalDiscount']

        so.save()
        request.user.branch.salesOrder.add(so)

        for jsonatc in salesOrder['atc']:

            atc = SOatc()
            print(jsonatc)
            atc.code = ATC.objects.get(pk=jsonatc['code'])
            atc.amountWithheld = jsonatc['amountWithheld']
            atc.salesOrder = so
            atc.save()
            request.user.branch.soatc.add(atc)

        for item in salesOrder['items']:
            soitemsmerch = SOItemsMerch()
            soitemsmerch.salesOrder = so
            soitemsmerch.merchInventory = MerchandiseInventory.objects.get(pk=item['code'])
            soitemsmerch.remaining = item['remaining']
            soitemsmerch.qty = item['quantity']
            soitemsmerch.cbm = item['cbm']
            soitemsmerch.vol = item['vol']
            soitemsmerch.pricePerCubic = item['pricePerCubic']
            soitemsmerch.totalCost = Decimal(item['totalCost'])

            print(soitemsmerch.totalCost)

            soitemsmerch.save()
            request.user.branch.soitemsMerch.add(soitemsmerch)

        for fee in salesOrder['otherFees']:
            f = SOOtherFees()
            f.salesOrder = so
            f.fee = fee['fee']
            f.description = fee['description']
            f.save()
            request.user.branch.soOtherFees.add(f)

        notify(request, 'New SO Request', so.code, '/so-nonapproved/', 1)
        sweetify.sweetalert(request, icon='success', title='Success!', persistent='Dismiss')
        return JsonResponse(0, safe=False)

class ExportSO(APIView):
    def get(self, request, pk):
        so = SalesOrder.objects.get(pk=pk)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="Sales Order {so.code}.xlsx"'


        wb = op.load_workbook('static/files/salesOrder.xlsx')
        ws = wb.active

        buyer = ws.cell(row=7, column=3)
        soCode = ws.cell(row=8, column=10)
        controlNo = ws.cell(row=6, column=10)
        date = ws.cell(row=7, column=11)

        gTotalPiecesQty = ws.cell(row=31, column=8)
        gTotalPiecesUnitPrice = ws.cell(row=31, column=9)
        gTotalAmount = ws.cell(row=30, column=10)
        gTotalVol = ws.cell(row=30, column=11)

        buyer.value=so.party.name
        soCode.value=so.code
        controlNo.value="Control No."
        date.value=now.today()
        
        gTotalPiecesQty.value=0
        gTotalPiecesUnitPrice.value=0
        gTotalAmount.value=0
        gTotalVol.value=0

        ctr = 0
        for i in so.soitemsmerch.all():

            invName = ws.cell(row=13+ctr, column=2)
            length = ws.cell(row=13+ctr, column=3)
            width = ws.cell(row=13+ctr, column=5)
            thick = ws.cell(row=13+ctr, column=7)
            qty = ws.cell(row=13+ctr, column=8)
            unitCost = ws.cell(row=13+ctr, column=9)
            amount = ws.cell(row=13+ctr, column=10)
            vol = ws.cell(row=13+ctr, column=11)

            invName.value=f"{i.merchInventory.name} {i.merchInventory.classification}"
            length.value=f"{round(i.merchInventory.length, 0)}"
            width.value=f"{round(i.merchInventory.width, 0)}"
            thick.value=f"{round(i.merchInventory.thickness, 0)}"
            qty.value=i.qty
            amount.value=round(i.totalCost, 2)
            vol.value=i.vol
            unitCost.value=(vol.value/qty.value)*i.pricePerCubic

            gTotalPiecesQty.value += qty.value
            gTotalPiecesUnitPrice.value += unitCost.value
            gTotalAmount.value += amount.value
            gTotalVol.value += vol.value

            ctr+=1

        wb.save(response)
        return response

class ExportSC(APIView):
    def get(self, request, pk):
        sc = SalesContract.objects.get(pk=pk)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="Sales Contract {sc.code}.xlsx"'

        wb = op.load_workbook('static/files/salesContract.xlsx')
        ws = wb.active

        buyer = ws.cell(row=7, column=3)
        scCode = ws.cell(row=6, column=11)
        date = ws.cell(row=7, column=12)

        buyer.value=sc.party.name
        scCode.value=sc.code
        date.value=now.today()

        ctr = 0
        for i in sc.scitemsmerch.all():
            invName = ws.cell(row=15+ctr, column=2)
            length = ws.cell(row=15+ctr, column=3)
            width = ws.cell(row=15+ctr, column=5)
            thick = ws.cell(row=15+ctr, column=7)
            qty = ws.cell(row=15+ctr, column=8)
            pricePerCubic = ws.cell(row=15+ctr, column=9)
            unitCost = ws.cell(row=15+ctr, column=10)

            invName.value=f"{i.merchInventory.name} {i.merchInventory.classification}"
            length.value=f"{round(i.merchInventory.length, 0)}"
            width.value=f"{round(i.merchInventory.width, 0)}"
            thick.value=f"{round(i.merchInventory.thickness, 0)}"
            qty.value=i.qty
            pricePerCubic.value=round(i.pricePerCubic, 2)

            ctr+=1
        
        wb.save(response)
        return response

class ExportDRLS(APIView):
    def get(self, request, pk):
        sc = SalesContract.objects.get(pk=pk)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="Loading Slip DR {sc.code}.xlsx"'

        wb = op.load_workbook('static/files/DR_LS.xlsx')
        ws = wb.active

        buyer = ws.cell(row=7, column=2)
        date = ws.cell(row=4, column=9)

        buyer.value=sc.party.name
        date.value=now.today()

        ctr = 0
        for i in sc.scitemsmerch.all():
            invName = ws.cell(row=12+ctr, column=1)
            length = ws.cell(row=12+ctr, column=2)
            width = ws.cell(row=12+ctr, column=3)
            thick = ws.cell(row=12+ctr, column=4)
            qty = ws.cell(row=12+ctr, column=5)

            invName.value=f"{i.merchInventory.name} {i.merchInventory.classification}"
            length.value=f"{round(i.merchInventory.length, 0)}"
            width.value=f"{round(i.merchInventory.width, 0)}"
            thick.value=f"{round(i.merchInventory.thickness, 0)}"
            qty.value=i.qty

            ctr+=1

        wb.save(response)
        return response

class ExportQuotationsSlip(APIView):
    def get(self, request, pk):
        qq = Quotations.objects.get(pk=pk)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="Quotations Slip {qq.code}.xlsx"'

        wb = op.load_workbook('static/files/QuotationSlip.xlsx')
        ws = wb.active

        buyer = ws.cell(row=7, column=2)
        date = ws.cell(row=4, column=9)
        code = ws.cell(row=2, column=6)

        buyer.value=qq.party.name
        code.value=qq.code
        date.value=now.today()

        ctr = 0

        for i in qq.qqitemsmerch.all():
            invName = ws.cell(row=12+ctr, column=1)
            length = ws.cell(row=12+ctr, column=2)
            width = ws.cell(row=12+ctr, column=3)
            thick = ws.cell(row=12+ctr, column=4)
            qty = ws.cell(row=12+ctr, column=5)

            invName.value=f"{i.merchInventory.name} {i.merchInventory.classification}"
            length.value=f"{round(i.merchInventory.length, 0)}"
            width.value=f"{round(i.merchInventory.width, 0)}"
            thick.value=f"{round(i.merchInventory.thickness, 0)}"
            qty.value=i.qty

            ctr+=1

        wb.save(response)
        return response